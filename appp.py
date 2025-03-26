# import streamlit as st
# import pandas as pd
# import requests
# import json
# import openpyxl
# import time
# import math

# st.title("Excel File Reader with Month and Year Filter")


# WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
# MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
# PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
# API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"


# if 'processed_df' not in st.session_state:
#     st.session_state.processed_df = None

# if 'total_count_df' not in st.session_state:
#     st.session_state.total_count_df = None

# # Function to get access token
# def GetAccesstoken():
#     auth_url = "https://iam.cloud.ibm.com/identity/token"
    
#     headers = {
#         "Content-Type": "application/x-www-form-urlencoded",
#         "Accept": "application/json"
#     }
    
#     data = {
#         "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
#         "apikey": API_KEY
#     }
#     response = requests.post(auth_url, headers=headers, data=data)
    
#     if response.status_code != 200:
#         st.write(f"Failed to get access token: {response.text}")
#         return None
#     else:
#         token_info = response.json()
#         return token_info['access_token']

# # Generate prompt for Watson API
# def generatePrompt(json_datas):
#     body = {
#         "input": f"""
#         Parse this JSON data, where each entry represents a task with 'Activity Name' and 'Finish_Month_Name' fields:
#         {json_datas}
# Count the number of occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)} for the selected year. 
# Return a JSON object where:
# - Keys are unique 'Activity Name' values.
# - Values are dictionaries with keys as 'Finish_Month_Name' from {', '.join(selected_months)} and values as the count of occurrences.
# - Include all months from {', '.join(selected_months)} for every 'Activity Name', setting the count to 0 if there are no occurrences in that month.
# Example output:
# {{
#     "Install Windows": {{"Mar": 2, "Apr": 1, "May": 0}},
#     "Paint Walls": {{"Mar": 1, "Apr": 3, "May": 0}}
# }}
# Return only the JSON object, no code, no explanation, just the formatted JSON.

#         """, 
#         "parameters": {
#             "decoding_method": "greedy",
#             "max_new_tokens": 8100,
#             "min_new_tokens": 0,
#             "stop_sequences": [";"],
#             "repetition_penalty": 1.05,
#             "temperature": 0.5
#         },
#         "model_id": MODEL_ID,
#         "project_id": PROJECT_ID
#     }
    
#     headers = {
#         "Accept": "application/json",
#         "Content-Type": "application/json",
#         "Authorization": f"Bearer {GetAccesstoken()}"
#     }
    
#     if not headers["Authorization"]:
#         return "Error: No valid access token."
    
#     response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
#     if response.status_code != 200:
#         st.write(f"Failed to generate prompt: {response.text}")
#         return "Error generating prompt"
#     # st.write(json_datas)
#     return response.json()['results'][0]['generated_text'].strip()

# # Function to create chunks and process data
# def createChunk(result_json, chunk_size=2000):
#     num_rows = len(result_json)
#     num_chunks = (num_rows + chunk_size - 1) // chunk_size  
#     all_chunks = {} 
#     temp = []

#     st.write(f"Filtered rows in JSON format (split into {chunk_size}-row chunks):")
 
#     for i in range(num_chunks):
#         start_idx = i * chunk_size
#         end_idx = min((i + 1) * chunk_size, num_rows)
#         chunk_data = result_json[start_idx:end_idx]

#         processed_data = json.loads(generatePrompt(chunk_data)) 
#         temp.append(processed_data)  

#         all_chunks[f"chunk_{i + 1}"] = chunk_data

    
#     table_data = []


#     all_months = set()

#     # Gather all unique months across all chunks
#     for chunk in temp:
#         for activity, months in chunk.items():
#             all_months.update(months.keys())  # Get all unique month names

#     # Sort months in order
#     all_months = sorted(all_months, key=lambda x: pd.to_datetime(x, format='%b').month)

#     # Now build the rows for the DataFrame
#     for chunk in temp:
#         for activity, months in chunk.items():
#             row = {'Activity Name': activity}
#             total_count = 0  # Initialize total count for each activity
#             for month in all_months:
#                 count = months.get(month, 0)  # Default to 0 if the month is not present for the activity
#                 row[month] = count
#                 total_count += count  # Add the month count to the total
#             row['Total Count'] = total_count  # Add total count for the activity
#             table_data.append(row)

#     # Convert the table_data into a DataFrame
#     df = pd.DataFrame(table_data)

#     # Display the DataFrame in Streamlit
#     st.write("Activity Counts Table by Month")
#     st.dataframe(df)


    


# # File upload and processing
# uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

# if uploaded_file is not None and st.session_state.processed_df is None:
#     workbook = openpyxl.load_workbook(uploaded_file)
#     sheet = workbook["TOWER 4 FINISHING."]
    
#     activity_col_idx = 5  

#     non_bold_rows = []
#     for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16):
#         cell = row[activity_col_idx]  
#         if cell.font and not cell.font.b:  
#             non_bold_rows.append(row_idx)

#     df = pd.read_excel(uploaded_file, sheet_name="TOWER 4 FINISHING.", skiprows=15)
    
#     df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
#                   'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
#                   'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
    
#     required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
#     df = df[required_columns]
    
#     df.index = df.index + 16
#     df = df.loc[df.index.isin(non_bold_rows)]

#     df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
#     df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)

#     df['Finish_Year'] = df['Finish'].dt.year
#     df['Finish_Month'] = df['Finish'].dt.month
#     df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')

#     st.session_state.processed_df = df

# # Process the filtered data
# if st.session_state.processed_df is not None:
#     df = st.session_state.processed_df
#     available_years = sorted(df['Finish_Year'].unique())
#     available_months = sorted(df['Finish_Month_Name'].unique())

#     selected_year = st.sidebar.selectbox('Select Year', available_years, index=available_years.index(df['Finish_Year'].max()))
#     selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

#     filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]

#     st.write(f"Filtered rows based on the selected months and year: {', '.join(selected_months)} {selected_year}")
#     st.write(filtered_df)
#     st.write(f"Number of rows: {len(filtered_df)}")

# #------------------------THIS IS JSON DATA---------------------

# #ithu ennoda json
#     result = filtered_df[['Activity Name', 'Finish_Month_Name']]

#     result_json = result.to_dict(orient='records')

#     # st.text(result_json)
#     df = pd.DataFrame(result_json)
    
#     mar_count = len(df[(df['Activity Name'] == 'EL-Third Fix ') & (df['Finish_Month_Name'] == 'Mar')])
    
#     st.write(f"Number of occurrences of 'Mar' for 'EL-Third Fix': {mar_count}")

# #------------------------THIS IS JSON DATA---------------------

#     activity_month_counts = pd.pivot_table(
#         filtered_df, 
#         values='Activity ID', 
#         index='Activity Name', 
#         columns='Finish_Month_Name', 
#         aggfunc='count', 
#         fill_value=0
#     )

#     activity_month_counts['Total Count'] = activity_month_counts.sum(axis=1)

#     st.write(f"Activity counts by month for {selected_year}:")
#     st.write(activity_month_counts)

# if st.button('Count The activity'):
#     createChunk(result_json)


# import streamlit as st
# import pandas as pd
# import requests
# import json
# import openpyxl
# import time

# st.title("Excel File Reader with Month and Year Filter")

# WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
# MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
# PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
# API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

# if 'processed_df' not in st.session_state:
#     st.session_state.processed_df = None

# # Function to get access token
# def get_access_token():
#     auth_url = "https://iam.cloud.ibm.com/identity/token"
#     headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
#     data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
#     for _ in range(3):
#         response = requests.post(auth_url, headers=headers, data=data)
#         if response.status_code == 200:
#             return response.json()['access_token']
#         time.sleep(1)
#     return None

# # Generate prompt for Watson API
# def generate_prompt(json_data, selected_months):
#     body = {
#         "input": f"""
# Given this JSON data where each entry has 'Activity Name' and 'Finish_Month_Name':
# {json.dumps(json_data, indent=2)}

# Perform the following task:
# 1. Count the occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)}.
# 2. Return a JSON object where:
#    - Keys are unique 'Activity Name' values.
#    - Values are objects with keys as the months {', '.join(selected_months)} and values as the count of occurrences.
#    - For each 'Activity Name', include all months from {', '.join(selected_months)}, setting the count to 0 if no occurrences exist.
# 3. Ensure the output is a valid JSON string with no additional text, comments, or explanations.

# Example input:
# [
#   {{"Activity Name": "Install Windows", "Finish_Month_Name": "Mar"}},
#   {{"Activity Name": "Install Windows", "Finish_Month_Name": "Mar"}},
#   {{"Activity Name": "Paint Walls", "Finish_Month_Name": "Apr"}}
# ]

# Expected output:
# {{
#   "Install Windows": {{"Mar": 2, "Apr": 0}},
#   "Paint Walls": {{"Mar": 0, "Apr": 1}}
# }}

# Return only the valid JSON object as a string.
# """,
#         "parameters": {
#             "decoding_method": "greedy",
#             "max_new_tokens": 8100,
#             "min_new_tokens": 0,
#             "stop_sequences": [],
#             "repetition_penalty": 1.0,
#             "temperature": 0.1
#         },
#         "model_id": MODEL_ID,
#         "project_id": PROJECT_ID
#     }
#     headers = {"Accept": "application/json", "Content-Type": "application/json", "Authorization": f"Bearer {get_access_token()}"}
#     if not headers["Authorization"]:
#         return None
    
#     try:
#         response = requests.post(WATSONX_API_URL, headers=headers, json=body)
#         response.raise_for_status()
#         raw_response = response.json()['results'][0]['generated_text'].strip()
#         json.loads(raw_response)  # Validate JSON
#         return raw_response
#     except (requests.RequestException, json.JSONDecodeError, KeyError):
#         return None

# # Process data and display only tables
# def process_and_display_data(json_data, selected_months):
#     # Watsonx API result (or fallback to pandas if API fails)
#     result = generate_prompt(json_data, selected_months)
#     if not result:
#         df = pd.DataFrame(json_data)
#         counts = pd.pivot_table(df, index='Activity Name', columns='Finish_Month_Name', 
#                                 aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#         result = counts.to_json(orient='index')
    
#     try:
#         activity_counts = json.loads(result)
#     except json.JSONDecodeError:
#         df = pd.DataFrame(json_data)
#         counts = pd.pivot_table(df, index='Activity Name', columns='Finish_Month_Name', 
#                                 aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#         activity_counts = json.loads(counts.to_json(orient='index'))

#     # Convert API result to DataFrame
#     df_data = []
#     for activity, months in activity_counts.items():
#         row = {'Activity Name': activity}
#         row.update(months)
#         row['Total Count'] = sum(months.values())
#         df_data.append(row)
    
#     result_df = pd.DataFrame(df_data).set_index('Activity Name')
#     st.write("Activity Counts by Month (Watsonx API):")
#     st.dataframe(result_df)

#     # Pandas validation table
#     pandas_df = pd.DataFrame(json_data)
#     validation_counts = pd.pivot_table(pandas_df, index='Activity Name', columns='Finish_Month_Name', 
#                                        aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#     validation_counts['Total Count'] = validation_counts.sum(axis=1)
#     st.write("Activity Counts by Month (Pandas Validation):")
#     st.dataframe(validation_counts)

# # File upload and processing
# uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

# if uploaded_file is not None and st.session_state.processed_df is None:
#     try:
#         workbook = openpyxl.load_workbook(uploaded_file)
#         sheet_names = workbook.sheetnames
#         if len(sheet_names) > 1:
#             selected_sheet = st.selectbox("Select a sheet", sheet_names)
#         else:
#             selected_sheet = sheet_names[0]
        
#         sheet = workbook[selected_sheet]
#         activity_col_idx = 5
#         non_bold_rows = [row_idx for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16) 
#                          if row[activity_col_idx].font and not row[activity_col_idx].font.b]

#         df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, skiprows=15)
#         if len(df.columns) >= 16:
#             df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
#                           'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
#                           'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
#         required_columns = ['Activity Name', 'Finish']
#         df = df[required_columns]
#         df.index = df.index + 16
#         df = df.loc[df.index.isin(non_bold_rows)]

#         df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)
#         df['Finish_Year'] = df['Finish'].dt.year
#         df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')
#         st.session_state.processed_df = df
#     except Exception as e:
#         st.error(f"Error processing file: {str(e)}")

# # Process filtered data
# if st.session_state.processed_df is not None:
#     df = st.session_state.processed_df
#     available_years = sorted(df['Finish_Year'].dropna().unique())
#     available_months = sorted(df['Finish_Month_Name'].dropna().unique())

#     selected_year = st.sidebar.selectbox('Select Year', available_years, index=len(available_years)-1)
#     selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

#     filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]
#     result_json = filtered_df[['Activity Name', 'Finish_Month_Name']].to_dict(orient='records')
    
#     if st.button('Count Activities'):
#         process_and_display_data(result_json, selected_months)


# app1.py
# import streamlit as st
# import pandas as pd
# import requests
# import json
# import openpyxl
# import time

# # Constants
# WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
# MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
# PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
# API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

# # Function to get access token
# def get_access_token():
#     auth_url = "https://iam.cloud.ibm.com/identity/token"
#     headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
#     data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
#     for _ in range(3):
#         response = requests.post(auth_url, headers=headers, data=data)
#         if response.status_code == 200:
#             return response.json()['access_token']
#         time.sleep(1)
#     return None

# # Generate prompt for Watson API
# def generate_prompt(json_data, selected_months):
#     body = {
#         "input": f"""
# Given this JSON data where each entry has 'Activity Name' and 'Finish_Month_Name':
# {json.dumps(json_data, indent=2)}

# Perform the following task:
# 1. Count the occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)}.
# 2. Return a JSON object where:
#    - Keys are unique 'Activity Name' values.
#    - Values are objects with keys as the months {', '.join(selected_months)} and values as the count of occurrences.
#    - For each 'Activity Name', include all months from {', '.join(selected_months)}, setting the count to 0 if no occurrences exist.
# 3. Ensure the output is a valid JSON string with no additional text, comments, or explanations.

# Example input:
# [
#   {{"Activity Name": "Install Windows", "Finish_Month_Name": "Mar"}},
#   {{"Activity Name": "Install Windows", "Finish_Month_Name": "Mar"}},
#   {{"Activity Name": "Paint Walls", "Finish_Month_Name": "Apr"}}
# ]

# Expected output:
# {{
#   "Install Windows": {{"Mar": 2, "Apr": 0}},
#   "Paint Walls": {{"Mar": 0, "Apr": 1}}
# }}

# Return only the valid JSON object as a string.
# """,
#         "parameters": {
#             "decoding_method": "greedy",
#             "max_new_tokens": 8100,
#             "min_new_tokens": 0,
#             "stop_sequences": [],
#             "repetition_penalty": 1.0,
#             "temperature": 0.1
#         },
#         "model_id": MODEL_ID,
#         "project_id": PROJECT_ID
#     }
#     headers = {"Accept": "application/json", "Content-Type": "application/json", "Authorization": f"Bearer {get_access_token()}"}
#     if not headers["Authorization"]:
#         return None
    
#     try:
#         response = requests.post(WATSONX_API_URL, headers=headers, json=body)
#         response.raise_for_status()
#         raw_response = response.json()['results'][0]['generated_text'].strip()
#         json.loads(raw_response)  # Validate JSON
#         return raw_response
#     except (requests.RequestException, json.JSONDecodeError, KeyError):
#         return None

# # Process data and display tables
# def process_and_display_data(json_data, selected_months):
#     result = generate_prompt(json_data, selected_months)
#     if not result:
#         df = pd.DataFrame(json_data)
#         counts = pd.pivot_table(df, index='Activity Name', columns='Finish_Month_Name', 
#                                 aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#         result = counts.to_json(orient='index')
    
#     try:
#         activity_counts = json.loads(result)
#     except json.JSONDecodeError:
#         df = pd.DataFrame(json_data)
#         counts = pd.pivot_table(df, index='Activity Name', columns='Finish_Month_Name', 
#                                 aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#         activity_counts = json.loads(counts.to_json(orient='index'))

#     df_data = []
#     for activity, months in activity_counts.items():
#         row = {'Activity Name': activity}
#         row.update(months)
#         row['Total Count'] = sum(months.values())
#         df_data.append(row)
    
#     result_df = pd.DataFrame(df_data).set_index('Activity Name')
#     st.write("Activity Counts by Month (Watsonx API):")
#     st.dataframe(result_df)

#     pandas_df = pd.DataFrame(json_data)
#     validation_counts = pd.pivot_table(pandas_df, index='Activity Name', columns='Finish_Month_Name', 
#                                        aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#     validation_counts['Total Count'] = validation_counts.sum(axis=1)
#     st.write("Activity Counts by Month (Pandas Validation):")
#     st.dataframe(validation_counts)

# def main():
#     """Main function for Application 1."""
#     st.title("Excel File Reader with Month and Year Filter")

#     if 'processed_df' not in st.session_state:
#         st.session_state.processed_df = None

#     # File upload and processing
#     uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

#     if uploaded_file is not None and st.session_state.processed_df is None:
#         try:
#             workbook = openpyxl.load_workbook(uploaded_file)
#             sheet_names = workbook.sheetnames
#             if len(sheet_names) > 1:
#                 selected_sheet = st.selectbox("Select a sheet", sheet_names)
#             else:
#                 selected_sheet = sheet_names[0]
            
#             sheet = workbook[selected_sheet]
#             activity_col_idx = 5
#             non_bold_rows = [row_idx for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16) 
#                              if row[activity_col_idx].font and not row[activity_col_idx].font.b]

#             df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)
#             if len(df.columns) >= 16:
#                 df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
#                               'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
#                               'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
#             required_columns = ['Activity Name', 'Finish']
#             df = df[required_columns]
#             df.index = df.index + 16
#             df = df.loc[df.index.isin(non_bold_rows)]

#             df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)
#             df['Finish_Year'] = df['Finish'].dt.year
#             df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')
#             st.session_state.processed_df = df
#         except Exception as e:
#             st.error(f"Error processing file: {str(e)}")

#     # Process filtered data
#     if st.session_state.processed_df is not None:
#         df = st.session_state.processed_df
#         available_years = sorted(df['Finish_Year'].dropna().unique())
#         available_months = sorted(df['Finish_Month_Name'].dropna().unique())

#         selected_year = st.sidebar.selectbox('Select Year', available_years, index=len(available_years)-1)
#         selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

#         filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]
#         result_json = filtered_df[['Activity Name', 'Finish_Month_Name']].to_dict(orient='records')
        
#         if st.button('Count Activities'):
#             process_and_display_data(result_json, selected_months)

# if __name__ == "__main__":
#     main()

# import streamlit as st
# import pandas as pd
# import requests
# import json
# import openpyxl
# import time
# from datetime import datetime

# # Constants remain the same
# WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
# MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
# PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
# API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

# # get_access_token() remains the same
# def get_access_token():
#     auth_url = "https://iam.cloud.ibm.com/identity/token"
#     headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
#     data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
#     for _ in range(3):
#         response = requests.post(auth_url, headers=headers, data=data)
#         if response.status_code == 200:
#             return response.json()['access_token']
#         time.sleep(1)
#     return None

# # generate_prompt() remains the same
# def generate_prompt(json_data, selected_months):
#     body = {
#         "input": f"""
# Given this JSON data where each entry has 'Activity Name' and 'Finish_Month_Name':
# {json.dumps(json_data, indent=2)}

# Perform the following task:
# 1. Count the occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)}.
# 2. Return a JSON object where:
#    - Keys are unique 'Activity Name' values.
#    - Values are objects with keys as the months {', '.join(selected_months)} and values as the count of occurrences.
#    - For each 'Activity Name', include all months from {', '.join(selected_months)}, setting the count to 0 if no occurrences exist.
# 3. Ensure the output is a valid JSON string with no additional text, comments, or explanations.

# Return only the valid JSON object as a string.
# """,
#         "parameters": {
#             "decoding_method": "greedy",
#             "max_new_tokens": 8100,
#             "min_new_tokens": 0,
#             "stop_sequences": [],
#             "repetition_penalty": 1.0,
#             "temperature": 0.1
#         },
#         "model_id": MODEL_ID,
#         "project_id": PROJECT_ID
#     }
#     headers = {"Accept": "application/json", "Content-Type": "application/json", "Authorization": f"Bearer {get_access_token()}"}
#     if not headers["Authorization"]:
#         return None
    
#     try:
#         response = requests.post(WATSONX_API_URL, headers=headers, json=body)
#         response.raise_for_status()
#         raw_response = str(response.json()['results'][0]['generated_text']).strip()
#         json.loads(raw_response)
#         return raw_response
#     except (requests.RequestException, json.JSONDecodeError, KeyError) as e:
#         st.error(f"Error in Watson API call: {str(e)}")
#         return None

# # process_and_display_data() remains the same
# def process_and_display_data(json_data, selected_months):
#     result = generate_prompt(json_data, selected_months)
    
#     if not result:
#         df = pd.DataFrame(json_data)
#         counts = pd.pivot_table(df, index='Activity Name', columns='Finish_Month_Name', 
#                                aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#         result = counts.to_json(orient='index')
    
#     try:
#         activity_counts = json.loads(result)
#     except json.JSONDecodeError:
#         st.warning("Invalid JSON from API, falling back to pandas calculation")
#         df = pd.DataFrame(json_data)
#         counts = pd.pivot_table(df, index='Activity Name', columns='Finish_Month_Name', 
#                                aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#         activity_counts = json.loads(counts.to_json(orient='index'))

#     df_data = []
#     for activity, months in activity_counts.items():
#         row = {'Activity Name': activity}
#         row.update(months)
#         row['Total Count'] = sum(months.values())
#         df_data.append(row)
    
#     result_df = pd.DataFrame(df_data).set_index('Activity Name')
#     st.write("Activity Counts by Month (Watsonx API):")
#     st.dataframe(result_df)

#     pandas_df = pd.DataFrame(json_data)
#     validation_counts = pd.pivot_table(pandas_df, index='Activity Name', columns='Finish_Month_Name', 
#                                       aggfunc='size', fill_value=0).reindex(columns=selected_months, fill_value=0)
#     validation_counts['Total Count'] = validation_counts.sum(axis=1)
#     st.write("Activity Counts by Month (Pandas Validation):")
#     st.dataframe(validation_counts)

# # Modified process_excel_file function
# def process_excel_file(uploaded_file, selected_sheet):
#     try:
#         workbook = openpyxl.load_workbook(uploaded_file)
#         sheet = workbook[selected_sheet]
        
#         # Read the Excel file without assuming column names initially
#         df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)
        
#         # Convert column names to lowercase for consistent matching
#         df.columns = [str(col).strip().lower() for col in df.columns]
        
#         # Define possible column names
#         possible_activity_cols = ['activity name', 'activity', 'task name', 'task']
#         possible_finish_cols = ['finish', 'end', 'completion date', 'finish date']
        
#         # Find matching columns
#         activity_col = next((col for col in df.columns if col in possible_activity_cols), None)
#         finish_col = next((col for col in df.columns if col in possible_finish_cols), None)
        
#         # If columns not found by name, try position-based fallback with safety check
#         if not activity_col or not finish_col:
#             if len(df.columns) >= 15:  # Check if enough columns exist
#                 activity_col = df.columns[5]  # 6th column
#                 finish_col = df.columns[14]  # 15th column
#             else:
#                 raise ValueError("Required columns not found and file has insufficient columns")
        
#         # Get index for non-bold row filtering
#         activity_col_idx = list(df.columns).index(activity_col)
#         non_bold_rows = [row_idx for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=len(df.columns)), start=1) 
#                         if row[activity_col_idx].font and not row[activity_col_idx].font.b]
        
#         # Select only required columns
#         df = df[[activity_col, finish_col]]
#         df.columns = ['Activity Name', 'Finish']
        
#         # Apply non-bold filter if applicable
#         if non_bold_rows:
#             df.index = df.index + 1
#             df = df.loc[df.index.isin(non_bold_rows)]
        
#         # Process dates
#         df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)
#         df['Finish_Year'] = df['Finish'].dt.year
#         df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')
        
#         return df.dropna(subset=['Finish'])
    
#     except Exception as e:
#         st.error(f"Error processing file: {str(e)}")
#         return None

# # Modified main function
# def main():
#     st.title("Excel File Reader with Month and Year Filter")

#     if 'processed_df' not in st.session_state:
#         st.session_state.processed_df = None
#     if 'selected_sheet' not in st.session_state:
#         st.session_state.selected_sheet = None

#     uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

#     if uploaded_file is not None:
#         workbook = openpyxl.load_workbook(uploaded_file)
#         sheet_names = workbook.sheetnames
        
#         if len(sheet_names) > 1:
#             st.session_state.selected_sheet = st.selectbox("Select a sheet", sheet_names)
#             if st.button("Confirm Sheet Selection"):
#                 st.session_state.processed_df = process_excel_file(uploaded_file, st.session_state.selected_sheet)
#         else:
#             st.session_state.selected_sheet = sheet_names[0]
#             st.session_state.processed_df = process_excel_file(uploaded_file, st.session_state.selected_sheet)

#     if st.session_state.processed_df is not None:
#         df = st.session_state.processed_df
#         available_years = sorted(df['Finish_Year'].dropna().unique())
#         available_months = sorted(df['Finish_Month_Name'].dropna().unique())

#         if not available_years or not available_months:
#             st.warning("No valid date data found in the 'Finish' column")
#             return

#         selected_year = st.sidebar.selectbox('Select Year', available_years, index=len(available_years)-1)
#         selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

#         filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]
#         result_json = filtered_df[['Activity Name', 'Finish_Month_Name']].to_dict(orient='records')
        
#         if st.button('Count Activities'):
#             process_and_display_data(result_json, selected_months)

# if __name__ == "__main__":
#     main()


# appp.py
import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math

# Constants
WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

# Function to get access token
def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']

# Generate prompt for Watson API
def generatePrompt(json_datas, selected_months):
    body = {
        "input": f"""
        Parse this JSON data, where each entry represents a task with 'Activity Name' and 'Finish_Month_Name' fields:
        {json.dumps(json_datas, indent=2)}
Count the number of occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)} for the selected year. 
Return a JSON object where:
- Keys are unique 'Activity Name' values.
- Values are dictionaries with keys as 'Finish_Month_Name' from {', '.join(selected_months)} and values as the count of occurrences.
- Include all months from {', '.join(selected_months)} for every 'Activity Name', setting the count to 0 if there are no occurrences in that month.
Example output:
{{
    "Install Windows": {{"Mar": 2, "Apr": 1, "May": 0}},
    "Paint Walls": {{"Mar": 1, "Apr": 3, "May": 0}}
}}
Return only the JSON object, no code, no explanation, just the formatted JSON.
        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    return response.json()['results'][0]['generated_text'].strip()

# Function to create chunks and process data
def createChunk(result_json, selected_months, chunk_size=2000):
    num_rows = len(result_json)
    num_chunks = (num_rows + chunk_size - 1) // chunk_size  
    all_chunks = {} 
    temp = []

    st.write(f"Filtered rows in JSON format (split into {chunk_size}-row chunks):")
 
    for i in range(num_chunks):
        start_idx = i * chunk_size
        end_idx = min((i + 1) * chunk_size, num_rows)
        chunk_data = result_json[start_idx:end_idx]

        processed_data = json.loads(generatePrompt(chunk_data, selected_months)) 
        temp.append(processed_data)  

        all_chunks[f"chunk_{i + 1}"] = chunk_data

    table_data = []
    all_months = set()

    # Gather all unique months across all chunks
    for chunk in temp:
        for activity, months in chunk.items():
            all_months.update(months.keys())  # Get all unique month names

    # Sort months in order
    all_months = sorted(all_months, key=lambda x: pd.to_datetime(x, format='%b').month)

    # Build the rows for the DataFrame
    for chunk in temp:
        for activity, months in chunk.items():
            row = {'Activity Name': activity}
            total_count = 0  # Initialize total count for each activity
            for month in all_months:
                count = months.get(month, 0)  # Default to 0 if month not present
                row[month] = count
                total_count += count  # Add to total
            row['Total Count'] = total_count  # Add total count
            table_data.append(row)

    # Convert to DataFrame
    df = pd.DataFrame(table_data)

    # Display in Streamlit
    st.write("Activity Counts Table by Month")
    st.dataframe(df)

def main():
    st.title("Excel File Reader with Month and Year Filter")

    if 'processed_df' not in st.session_state:
        st.session_state.processed_df = None

    if 'total_count_df' not in st.session_state:
        st.session_state.total_count_df = None

    # File upload and processing
    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

    if uploaded_file is not None and st.session_state.processed_df is None:
        workbook = openpyxl.load_workbook(uploaded_file)
        if "TOWER 4 FINISHING." in workbook.sheetnames:
            sheet = workbook["TOWER 4 FINISHING."]
            activity_col_idx = 5  

            non_bold_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16):
                cell = row[activity_col_idx]  
                if cell.font and not cell.font.b:  
                    non_bold_rows.append(row_idx)

            df = pd.read_excel(uploaded_file, sheet_name="TOWER 4 FINISHING.", skiprows=15)
            
            df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
                        'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                        'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
            
            required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
            df = df[required_columns]
            
            df.index = df.index + 16
            df = df.loc[df.index.isin(non_bold_rows)]

            df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
            df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)

            df['Finish_Year'] = df['Finish'].dt.year
            df['Finish_Month'] = df['Finish'].dt.month
            df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')

            st.session_state.processed_df = df
        else:
            activity_col_idx = 5 
            sheet = workbook["M7 T5"]
            non_bold_rows = []
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16):
                cell = row[activity_col_idx]  
                if cell.font and not cell.font.b:  
                    non_bold_rows.append(row_idx)
                    
            df = pd.read_excel(uploaded_file, sheet_name="M7 T5", skiprows=1)
            
            expected_columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
                       'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                       'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
            
            actual_cols = len(df.columns)
            if actual_cols >= len(expected_columns):
                df.columns = expected_columns + [f'Extra_{i}' for i in range(actual_cols - len(expected_columns))]
            else:
                df.columns = expected_columns[:actual_cols]

            required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
            available_columns = [col for col in required_columns if col in df.columns]
            df = df[available_columns]

            df.index = df.index + 16
            df = df.loc[df.index.isin(non_bold_rows)]

            if 'Start' in df.columns:
                df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
            if 'Finish' in df.columns:
                df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)
                df['Finish_Year'] = df['Finish'].dt.year
                df['Finish_Month'] = df['Finish'].dt.month
                df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')

            st.session_state.processed_df = df

    # Process the filtered data
    if st.session_state.processed_df is not None:
        df = st.session_state.processed_df
        available_years = sorted(df['Finish_Year'].dropna().unique())  # Drop NaN for safety
        available_months = sorted(df['Finish_Month_Name'].dropna().unique())

        selected_year = st.sidebar.selectbox('Select Year', available_years, index=available_years.index(df['Finish_Year'].max()))
        selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

        filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]

        st.write(f"Filtered rows based on the selected months and year: {', '.join(selected_months)} {selected_year}")
        st.write(filtered_df)
        st.write(f"Number of rows: {len(filtered_df)}")

        # JSON data processing
        result = filtered_df[['Activity Name', 'Finish_Month_Name']]
        result_json = result.to_dict(orient='records')

        df_json = pd.DataFrame(result_json)
        mar_count = len(df_json[(df_json['Activity Name'] == 'EL-Third Fix ') & (df_json['Finish_Month_Name'] == 'Mar')])
        st.write(f"Number of occurrences of 'Mar' for 'EL-Third Fix': {mar_count}")

        # Pivot table for activity counts
        activity_month_counts = pd.pivot_table(
            filtered_df, 
            values='Activity ID', 
            index='Activity Name', 
            columns='Finish_Month_Name', 
            aggfunc='count', 
            fill_value=0
        )
        activity_month_counts['Total Count'] = activity_month_counts.sum(axis=1)

        st.write(f"Activity counts by month for {selected_year}:")
        st.write(activity_month_counts)

        if st.button('Count The Activity'):
            createChunk(result_json, selected_months)

if __name__ == "__main__":
    main()
